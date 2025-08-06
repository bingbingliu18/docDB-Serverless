#!/usr/bin/env python3
"""
DocumentDB Serverless vs On-Demand Cost Comparison Tool
"""

import boto3
import traceback
import math
import json
import os
from datetime import datetime, timedelta
import time
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import logging
import sys
from operator import itemgetter
import matplotlib.pyplot as plt
import numpy as np
import concurrent.futures
from botocore.exceptions import ClientError

priceList = []
counter = 0

IMG_WIDTH = 600
IMG_HEIGHT = 400

# List of supported DocumentDB regions
region_list = ['us-east-1','ap-northeast-1','us-east-2','us-west-1','us-west-2','ap-east-1','ap-south-1','ap-southeast-1','ap-northeast-2','ap-southeast-2','ca-central-1','eu-central-1','eu-west-1','eu-west-2','eu-west-3','eu-north-1','me-south-1','sa-east-1']

print("Please select a region by entering the corresponding number:")
for i, rg in enumerate(region_list, start=1):
    print(f"{i}. {rg}")
user_input = input(f"Enter your choice (1-{len(region_list)}): ")
if user_input.isdigit() and 1 <= int(user_input) <= len(region_list):
    docdb_region = region_list[int(user_input) - 1]
    print(f"You selected: {docdb_region}")
else:
    print("Invalid input. Please try again.")
    sys.exit(1)

os.environ['AWS_DEFAULT_REGION'] = docdb_region

# Create a new Excel workbook
myworkbook = openpyxl.Workbook()
myworksheet = myworkbook.active
myworksheet.title = "Summary"

# Get current date and time for log filename
log_filename = f"docdb_cost_comparison_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# Configure logging
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

def aws_region_to_location(region):
    """
    Convert AWS region code to location name used in pricing API
    """
    region_to_location_map = {
        "us-east-1": "US East (N. Virginia)",
        "us-east-2": "US East (Ohio)",
        "us-west-1": "US West (N. California)",
        "us-west-2": "US West (Oregon)",
        "ap-east-1": "Asia Pacific (Hong Kong)",
        "ap-south-1": "Asia Pacific (Mumbai)",
        "ap-northeast-1": "Asia Pacific (Tokyo)",
        "ap-northeast-2": "Asia Pacific (Seoul)",
        "ap-southeast-1": "Asia Pacific (Singapore)",
        "ap-southeast-2": "Asia Pacific (Sydney)",
        "ca-central-1": "Canada (Central)",
        "eu-central-1": "EU (Frankfurt)",
        "eu-west-1": "EU (Ireland)",
        "eu-west-2": "EU (London)",
        "eu-west-3": "EU (Paris)",
        "eu-north-1": "EU (Stockholm)",
        "me-south-1": "Middle East (Bahrain)",
        "sa-east-1": "South America (São Paulo)"
    }

    if region in region_to_location_map:
        return region_to_location_map[region]
    else:
        return "Unknown location"

def get_all_pages(pricing_client, service_code, filters, max_retries=5):
    """
    Get all results from AWS Pricing API with pagination
    """
    all_results = []
    next_token = None
    retries = 0
    
    while True:
        try:
            params = {
                'ServiceCode': service_code,
                'Filters': filters
            }
            if next_token:
                params['NextToken'] = next_token
            
            response = pricing_client.get_products(**params)
            all_results.extend(response['PriceList'])
            
            if 'NextToken' in response:
                next_token = response['NextToken']
            else:
                break
                
            time.sleep(0.5)
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'Throttling':
                retries += 1
                if retries > max_retries:
                    logging.error(f"Reached maximum retry attempts: {e}")
                    break
                wait_time = 2 ** retries
                logging.warning(f"Request throttled, retrying in {wait_time} seconds ({retries}/{max_retries})")
                time.sleep(wait_time)
            else:
                logging.error(f"API request error: {e}")
                break
    
    logging.info(f"Retrieved {len(all_results)} pricing records")
    return all_results

def pricing_get_product(engine, instance_class=None):
    """
    Get DocumentDB product pricing information from price list
    """
    global priceList
    logging.info(f"Searching for price: instance_class={instance_class}, engine={engine}")
    for price in priceList:
        if (price['product']['productFamily'] == 'Database Instance' and 
            price['product']['attributes']['databaseEngine'].lower() == engine.lower() and 
            price['product']['attributes']['instanceType'] == instance_class):
            return price
    logging.error(f'param: engine={engine}, instanceType={instance_class}')
    logging.info(f'priceList Dump: {json.dumps(priceList, indent=2)}')
    raise Exception('price not found')

def get_docdb_dcu_price(docdb_region):
    """
    Get DocumentDB Serverless DCU(IO-Optimized) unit price
    Enhanced real-time query based on docdb_dcu_price_query.py logic
    """
    pricing_client = boto3.client('pricing', region_name='us-east-1')
    location = aws_region_to_location(docdb_region)
    
    logging.info(f"Querying real-time DCU price for {docdb_region} ({location})...")
    
    # Set filter conditions to search for DocumentDB pricing
    filters = [
        {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location}
    ]
    
    try:
        # Get DocumentDB pricing data
        products = get_all_pages(pricing_client, 'AmazonDocDB', filters)
        price_list = [json.loads(product) for product in products]
        
        logging.info(f"Retrieved {len(price_list)} DocumentDB pricing records")
        
        # Search for Serverless-related pricing first
        serverless_prices = []
        
        for price in price_list:
            attributes = price['product']['attributes']
            
            # Check if this is a Serverless-related product
            usage_type = attributes.get('usagetype', '').lower()
            product_family = attributes.get('productFamily', '').lower()
            instance_type = attributes.get('instanceType', '').lower()
            description = attributes.get('description', '').lower()
            
            # Search for products containing serverless keywords
            if ('serverless' in usage_type or 
                'serverless' in product_family or 
                'serverless' in instance_type or
                'serverless' in description):
                
                # Extract pricing information
                terms = price.get('terms', {})
                on_demand = terms.get('OnDemand', {})
                
                for offer_term_code, offer_term_data in on_demand.items():
                    for price_dimension_key, price_dimension_data in offer_term_data['priceDimensions'].items():
                        price_per_unit = float(price_dimension_data['pricePerUnit']['USD'])
                        unit = price_dimension_data.get('unit', 'Unknown')
                        
                        serverless_prices.append({
                            'price': price_per_unit,
                            'unit': unit,
                            'usage_type': usage_type,
                            'product_family': product_family,
                            'description': price_dimension_data.get('description', 'N/A')
                        })
        
        # If found serverless prices, return the most appropriate one
        if serverless_prices:
            # Prioritize DCU-related units
            dcu_candidates = [p for p in serverless_prices if 'dcu' in p['unit'].lower() or 'capacity' in p['description'].lower()]
            
            if dcu_candidates:
                best_price = dcu_candidates[0]['price']
                logging.info(f"Found Serverless DCU price: ${best_price} per {dcu_candidates[0]['unit']}")
                return best_price
            else:
                # Use the first serverless price found
                best_price = serverless_prices[0]['price']
                logging.info(f"Found Serverless price: ${best_price} per {serverless_prices[0]['unit']}")
                return best_price
        
        # If no serverless pricing found, search for DCU-related pricing
        logging.warning("No DocumentDB Serverless pricing found, searching for DCU-related pricing...")
        
        for price in price_list:
            attributes = price['product']['attributes']
            
            # Check if this is a DCU-related product
            usage_type = attributes.get('usagetype', '').lower()
            product_family = attributes.get('productFamily', '').lower()
            description = attributes.get('description', '').lower()
            
            # Search for products containing DCU or capacity keywords
            if ('dcu' in usage_type or 
                'dcu' in description or
                'capacity' in description or
                'io-optimized' in description):
                
                # Extract pricing information
                terms = price.get('terms', {})
                on_demand = terms.get('OnDemand', {})
                
                for offer_term_code, offer_term_data in on_demand.items():
                    for price_dimension_key, price_dimension_data in offer_term_data['priceDimensions'].items():
                        price_per_unit = float(price_dimension_data['pricePerUnit']['USD'])
                        unit = price_dimension_data.get('unit', 'Unknown')
                        
                        # Prioritize IO-Optimized DCU pricing
                        if 'io-optimized' in description or 'dcu' in unit.lower():
                            logging.info(f"Found DCU-related price: ${price_per_unit} per {unit}")
                            return price_per_unit
        
        # If no specific pricing found, use region-based estimated pricing
        logging.warning("No DocumentDB Serverless DCU pricing found in API, using region-based estimates")
        
        # Region-specific estimated DCU pricing
        region_dcu_pricing = {
            'us-east-1': 0.0822,      # Standard DCU price
            'us-east-2': 0.0822,
            'us-west-1': 0.0904,
            'us-west-2': 0.0822,
            'ap-northeast-1': 0.0986,
            'ap-northeast-2': 0.0986,
            'ap-southeast-1': 0.0986,
            'ap-southeast-2': 0.0986,
            'ap-south-1': 0.0904,
            'ap-east-1': 0.1068,
            'eu-central-1': 0.0986,
            'eu-west-1': 0.0904,
            'eu-west-2': 0.0986,
            'eu-west-3': 0.0986,
            'eu-north-1': 0.0822,
            'ca-central-1': 0.0904,
            'me-south-1': 0.1068,
            'sa-east-1': 0.1150
        }
        
        estimated_price = region_dcu_pricing.get(docdb_region, 0.0905)  # Default IO-optimized price
        logging.info(f"Using estimated DCU price for {docdb_region}: ${estimated_price} per DCU-Hr")
        return estimated_price
        
    except Exception as e:
        logging.error(f"Error retrieving DCU price: {e}")
        # Return region-specific fallback price
        region_fallback_pricing = {
            'us-east-1': 0.0822,
            'us-east-2': 0.0822,
            'us-west-1': 0.0904,
            'us-west-2': 0.0822,
            'ap-northeast-1': 0.0986,
            'ap-southeast-1': 0.0986,
            'eu-west-1': 0.0904,
            'eu-central-1': 0.0986,
        }
        
        fallback_price = region_fallback_pricing.get(docdb_region, 0.0905)
        logging.info(f"Using fallback DCU price for {docdb_region}: ${fallback_price} per DCU-Hr")
        return fallback_price

def pricing_get_products_optimized(docdb_region):
    """
    Get DocumentDB pricing information for specified region
    """
    global priceList
    priceList = []
    
    pricing_client = boto3.client('pricing', region_name='us-east-1')
    location = aws_region_to_location(docdb_region)
    
    logging.info(f"Starting to retrieve DocumentDB ({docdb_region}) pricing")
    
    # Get Database Instance type pricing
    filters = [
        {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'Amazon DocumentDB'},
        {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
        {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Database Instance'}
    ]
    
    products = get_all_pages(pricing_client, 'AmazonDocDB', filters)
    for product in products:
        priceList.append(json.loads(product))
    
    logging.info(f"Completed DocumentDB pricing retrieval, collected {len(priceList)} pricing records")

def calculate_serverless_cost_estimate(vcpu, avg_cpu_util, min_cpu_util, max_cpu_util, dcu_price_per_hour, cpu_percent_data=None):
    """
    Calculate estimated DocumentDB Serverless cost using dual methods for a single instance
    Method 1: Simple average-based calculation
    Method 2: Detailed calculation with minimum DCU + elastic scaling cost
    """
    
    # Method 1: Simple average-based calculation
    # Calculate average DCU based on CPU utilization
    avg_dcu = math.ceil(avg_cpu_util / 100 * int(vcpu) * 4)
    serverless_cost_method1 = avg_dcu * 730 * dcu_price_per_hour
    
    logging.info(f"Method 1 - Simple calculation: vCPU={vcpu}, avg_cpu_util={avg_cpu_util}%, "
                f"calculated avg_dcu={avg_dcu}, "
                f"monthly cost per instance=${avg_dcu * 730 * dcu_price_per_hour:.2f}")
    
    # Method 2: Detailed calculation with minimum DCU + elastic scaling
    # Calculate minimum DCU baseline based on average and minimum CPU utilization
    # Method 2: Detailed calculation with minimum DCU + elastic scaling
    # Calculate minimum DCU baseline based on average and minimum CPU utilization
    min_dcu_baseline = math.ceil((avg_cpu_util + min_cpu_util) / 100 / 2 * int(vcpu) * 4)
    min_dcu_baseline = max(0.5, min_dcu_baseline)  # Ensure minimum 0.5 DCU
    
    # Base cost (minimum DCU)
    base_cost_total = min_dcu_baseline * dcu_price_per_hour * 730
    
    # Calculate elastic scaling cost (if detailed CPU data is available)
    elastic_cost_total = 0
    if cpu_percent_data and len(cpu_percent_data) > 0:
        # Calculate elastic scaling threshold
        elastic_threshold = (avg_cpu_util + min_cpu_util) / 2
        
        logging.info(f"Calculating elastic cost with threshold: {elastic_threshold}%")
        
        exceed_count = 0
        sum_exceed_cost = 0
        
        for cpu_value in cpu_percent_data:
            if cpu_value > elastic_threshold:
                # Calculate DCU required above baseline (using 4x multiplier)
                required_dcu = math.ceil(cpu_value / 100 * int(vcpu) * 4)
                additional_dcu = max(0, required_dcu - min_dcu_baseline)
                
                # Additional cost charged per minute
                additional_cost = additional_dcu * dcu_price_per_hour / 60
                sum_exceed_cost += additional_cost
                exceed_count += 1
        
        # Total elastic cost
        elastic_cost_total = sum_exceed_cost
        
        logging.info(f"Elastic scaling: {exceed_count} data points exceeded threshold, "
                    f"additional cost per instance=${sum_exceed_cost:.2f}")
    
    # Method 2 total cost
    serverless_cost_method2 = base_cost_total + elastic_cost_total
    
    logging.info(f"Method 2 - Detailed calculation: min_dcu_baseline={min_dcu_baseline}, "
                f"base_cost=${base_cost_total:.2f}, elastic_cost=${elastic_cost_total:.2f}, "
                f"total monthly cost=${serverless_cost_method2:.2f}")
    
    return {
        'method1_cost': round(serverless_cost_method1, 2),
        'method2_cost': round(serverless_cost_method2, 2),
        'min_dcu_baseline': min_dcu_baseline,
        'avg_dcu': avg_dcu
    }
def get_docdb_cpu_utilization(instance_id, region):
    """
    Get CPU utilization statistics for a given DocumentDB instance over the past month.
    Enhanced to include minute-level data for elastic cost calculation
    """
    cloudwatch_client = boto3.client('cloudwatch', region_name=region)

    response = cloudwatch_client.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'cpu_avg',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/DocDB',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31, 
                    'Stat': 'Average'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_min',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/DocDB',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,
                    'Stat': 'Minimum'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_max',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/DocDB',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,
                    'Stat': 'Maximum'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_percent',
                'Expression': 'IF(m1>0,m1)', 
                'Label': 'CPU Watch'
            },
            {
                'Id': 'm1',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/DocDB',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 60,  # 1-minute resolution for detailed analysis
                    'Stat': 'Maximum',
                    'Unit': 'Percent'
                },
                'ReturnData': False
            }
        ],
      
        StartTime=(datetime.utcnow() - timedelta(days=30)).isoformat() + 'Z',
        EndTime=datetime.utcnow().isoformat() + 'Z'
    )

    if response['MetricDataResults']:
        metrics = {}
        for mdr in response['MetricDataResults']:
            metrics[mdr['Id']] = {'timestamps': mdr['Timestamps'], 'values': mdr['Values']}
        return metrics
    else:
        return None

def get_docdb_storage_utilization(instance_id, region):
    """
    Get storage utilization for a given DocumentDB instance.
    """
    cloudwatch_client = boto3.client('cloudwatch', region_name=region)

    response = cloudwatch_client.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'storage_used',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/DocDB',
                        'MetricName': 'VolumeBytesUsed',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400, 
                    'Stat': 'Average'
                },
                'ReturnData': True
            }
        ],
      
        StartTime=(datetime.utcnow() - timedelta(days=7)).isoformat() + 'Z',
        EndTime=datetime.utcnow().isoformat() + 'Z'
    )

    if response['MetricDataResults']:
        metrics = {}
        for mdr in response['MetricDataResults']:
            metrics[mdr['Id']] = {'timestamps': mdr['Timestamps'], 'values': mdr['Values']}
        return metrics
    else:
        return None

def count_cpu_usage_distribution(cpu_usage_data):
    """
    Count CPU utilization distribution.
    """
    usage_ranges = [
        ('0% - 10%', 0, 10),
        ('10% - 20%', 10, 20),
        ('20% - 30%', 20, 30),
        ('30% - 50%', 30, 50),
        ('50% and above', 50, 100)
    ]

    usage_counts = [0] * len(usage_ranges)

    for usage in cpu_usage_data:
        for i, (_, min_range, max_range) in enumerate(usage_ranges):
            if min_range <= usage < max_range:
                usage_counts[i] += 1
                break

    result = [['CPU Usage Range', 'Percentage']]
    result.extend([[range_name, count] for range_name, count in zip(
        [range_name for range_name, _, _ in usage_ranges], usage_counts
    )])
    return result

def create_cpu_usage_distribution_chart(data, worksheet, workbook):
    """
    Create CPU utilization distribution pie chart and add it to Excel worksheet.
    """
    for row in data:
        worksheet.append(row)

    labels = [row[0] for row in data[1:] if float(row[1]) > 0]
    values = [float(row[1]) for row in data[1:] if float(row[1]) > 0]

    fig, ax = plt.subplots(figsize=(8, 8))
    patches, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%')
    
    for t in texts:
        t.set_size('smaller')
        t.set_color('black')

    for t in autotexts:
        t.set_size('smaller')
        t.set_color('white')

    ax.axis('equal')
    ax.set_title("DocumentDB Cluster Count by CPU Avg Utilization")

    plt.savefig("docdb_cpu_usage_pie.jpg", dpi=300)

    img = Image("docdb_cpu_usage_pie.jpg")
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    worksheet.add_image(img, "A16")

    workbook.save("docdb_cost_comparison_report.xlsx")

def create_cost_comparison_chart(data, worksheet, workbook):
    """
    Create cost comparison chart with dual calculation methods
    """
    id = 2
    worksheet.cell(row=1, column=12, value='instance')
    worksheet.cell(row=1, column=13, value='on-demand cost')
    worksheet.cell(row=1, column=14, value='serverless cost method1')
    worksheet.cell(row=1, column=15, value='serverless cost method2')
    worksheet.cell(row=1, column=16, value='cost savings method1')
    worksheet.cell(row=1, column=17, value='cost savings method2')
    
    cluster_cost_data = []
    
    for row in data:
        cells = row.split(',')
        # Select instance name, on-demand cost, serverless costs (both methods)
        selected_columns = [1, 13, 16, 17, 18, 19]  # instance_id, od_cost, serverless_method1, serverless_method2, savings_method1, savings_method2
        selected_cell = [cells[i] for i in selected_columns]
        
        for col, value in enumerate(selected_cell, start=1):
            if col > 1:  # All except instance name are numeric
                value = float(value)
            worksheet.cell(row=id, column=12 + col - 1, value=value)
        
        cluster_cost_data.extend(selected_cell)
        id = id + 1

    # Create cost comparison bar chart
    instances = []
    od_costs = []
    serverless_costs_method1 = []
    serverless_costs_method2 = []
    savings_method1 = []
    savings_method2 = []

    for i in range(0, len(cluster_cost_data), 6):
        instances.append(cluster_cost_data[i])
        od_costs.append(float(cluster_cost_data[i+1]))
        serverless_costs_method1.append(float(cluster_cost_data[i+2]))
        serverless_costs_method2.append(float(cluster_cost_data[i+3]))
        savings_method1.append(float(cluster_cost_data[i+4]))
        savings_method2.append(float(cluster_cost_data[i+5]))

    img = Image(create_cost_comparison_bar_chart(instances, od_costs, serverless_costs_method1, serverless_costs_method2, savings_method1, savings_method2))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT

    worksheet.add_image(img, "L16")
    workbook.save("docdb_cost_comparison_report.xlsx")

def create_cost_comparison_bar_chart(instances, od_costs, serverless_costs_method1, serverless_costs_method2, savings_method1, savings_method2):
    """
    Create cost comparison bar chart JPG image with dual calculation methods
    """
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # First chart: cost comparison - Method 1
    bar_width = 0.35
    x = np.arange(len(instances))
    
    bars1 = ax1.bar(x - bar_width/2, od_costs, bar_width, label='On-Demand', color='skyblue')
    bars2 = ax1.bar(x + bar_width/2, serverless_costs_method1, bar_width, label='Serverless (Method 1)', color='lightcoral')

    # Add value labels
    for i, (od, sl) in enumerate(zip(od_costs, serverless_costs_method1)):
        ax1.text(x[i] - bar_width/2, od, f'${od:.0f}', ha='center', va='bottom', fontsize=8)
        ax1.text(x[i] + bar_width/2, sl, f'${sl:.0f}', ha='center', va='bottom', fontsize=8)

    ax1.set_title("DocumentDB Cost Comparison: On-Demand vs Serverless (Method 1)", fontsize=12)
    ax1.set_xlabel("Instance", fontsize=10)
    ax1.set_ylabel("Monthly Cost (USD)", fontsize=10)
    ax1.set_xticks(x)
    ax1.set_xticklabels(instances, rotation=20, fontsize=8)
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Second chart: cost comparison - Method 2
    bars3 = ax2.bar(x - bar_width/2, od_costs, bar_width, label='On-Demand', color='skyblue')
    bars4 = ax2.bar(x + bar_width/2, serverless_costs_method2, bar_width, label='Serverless (Method 2)', color='orange')

    # Add value labels
    for i, (od, sl) in enumerate(zip(od_costs, serverless_costs_method2)):
        ax2.text(x[i] - bar_width/2, od, f'${od:.0f}', ha='center', va='bottom', fontsize=8)
        ax2.text(x[i] + bar_width/2, sl, f'${sl:.0f}', ha='center', va='bottom', fontsize=8)

    ax2.set_title("DocumentDB Cost Comparison: On-Demand vs Serverless (Method 2)", fontsize=12)
    ax2.set_xlabel("Instance", fontsize=10)
    ax2.set_ylabel("Monthly Cost (USD)", fontsize=10)
    ax2.set_xticks(x)
    ax2.set_xticklabels(instances, rotation=20, fontsize=8)
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    # Third chart: savings comparison - Method 1
    colors1 = ['green' if s > 0 else 'red' for s in savings_method1]
    bars5 = ax3.bar(instances, savings_method1, color=colors1, alpha=0.7)
    
    # Add value labels
    for i, s in enumerate(savings_method1):
        ax3.text(i, s, f'${s:.0f}', ha='center', va='bottom' if s > 0 else 'top', fontsize=8)
    
    ax3.set_title("Cost Savings - Method 1 (Positive = Serverless Cheaper)", fontsize=12)
    ax3.set_xlabel("Instance", fontsize=10)
    ax3.set_ylabel("Savings (USD)", fontsize=10)
    ax3.set_xticklabels(instances, rotation=20, fontsize=8)
    ax3.axhline(y=0, color='black', linestyle='-', alpha=0.3)
    ax3.grid(True, alpha=0.3)
    
    # Fourth chart: savings comparison - Method 2
    colors2 = ['green' if s > 0 else 'red' for s in savings_method2]
    bars6 = ax4.bar(instances, savings_method2, color=colors2, alpha=0.7)
    
    # Add value labels
    for i, s in enumerate(savings_method2):
        ax4.text(i, s, f'${s:.0f}', ha='center', va='bottom' if s > 0 else 'top', fontsize=8)
    
    ax4.set_title("Cost Savings - Method 2 (Positive = Serverless Cheaper)", fontsize=12)
    ax4.set_xlabel("Instance", fontsize=10)
    ax4.set_ylabel("Savings (USD)", fontsize=10)
    ax4.set_xticklabels(instances, rotation=20, fontsize=8)
    ax4.axhline(y=0, color='black', linestyle='-', alpha=0.3)
    ax4.grid(True, alpha=0.3)

    plt.tight_layout()
    
    file_name = "docdb_cost_comparison_chart_enhanced.jpg"
    plt.savefig(file_name, dpi=300, bbox_inches='tight')
    plt.close()
    return file_name

def update_progress(current_step, total_steps):
    """
    Update progress bar
    """
    progress = current_step / total_steps * 100
    bar_length = 30
    block = int(round(bar_length * progress / 100))
    text = "\rProcessing: [{0}] {1}%".format("#" * block + "-" * (bar_length - block), round(progress, 2))
    sys.stdout.write(text)
    sys.stdout.flush()
def process_instance(instance, instance_count, dcu_price_per_hour):
    """
    Process a single DocumentDB instance, get its resource usage and cost comparison information
    """
    global counter
    avg_cpu_list = []
    counter += 1
    update_progress(counter, instance_count)
    logging.info("-----------------------")
    
    instance_id = instance['DBInstanceIdentifier']
    instance_class = instance['DBInstanceClass']
    engine = instance['Engine']
    engine_version = instance['EngineVersion']
    status = instance['DBInstanceStatus']
    availability_zone = instance['AvailabilityZone']
    cluster_id = instance.get('DBClusterIdentifier', '')
    account_id = instance['DBInstanceArn'].split(':')[4]
    
    try:
        # Get pricing information
        product_json = pricing_get_product(engine="Amazon DocumentDB", instance_class=instance_class)
        vcpu = int(product_json['product']['attributes']['vcpu'])
        
        # Get OD unit price
        od_price_per_unit = 0
        for offer_term_code, offer_term_data in product_json['terms']["OnDemand"].items():
            for price_dimension_code, price_dimension_data in offer_term_data["priceDimensions"].items():
                od_price_per_unit = round(float(price_dimension_data["pricePerUnit"]["USD"]), 3)
                logging.info(f"OD Price per unit: {od_price_per_unit}")
                break
        
        # Get CPU utilization for this specific instance
        cpu_utils = get_docdb_cpu_utilization(instance_id, docdb_region)
        if cpu_utils and cpu_utils.get('cpu_avg') and cpu_utils.get('cpu_avg')['values']:
            avg_cpu_util = math.ceil(cpu_utils.get('cpu_avg')['values'][0])
            min_cpu_util = math.ceil(cpu_utils.get('cpu_min')['values'][0])
            max_cpu_util = math.ceil(cpu_utils.get('cpu_max')['values'][0])
        else:
            avg_cpu_util = 0
            min_cpu_util = 0
            max_cpu_util = 0
            logging.warning(f"Instance {instance_id} has no CPU utilization data")
        
        # Get storage utilization for this specific instance
        storage_utils = get_docdb_storage_utilization(instance_id, docdb_region)
        storage_used_gb = 0
        if storage_utils and storage_utils.get('storage_used') and storage_utils.get('storage_used')['values']:
            storage_used_bytes = storage_utils.get('storage_used')['values'][0]
            storage_used_gb = round(storage_used_bytes / (1024**3), 2)  # Convert to GB
        
        # Calculate On-Demand cost for single instance
        od_monthly_cost = round(730 * od_price_per_unit, 2)
        
        # Calculate Serverless cost using dual methods for single instance
        cpu_percent_values = None
        if cpu_utils and cpu_utils.get('cpu_percent') and cpu_utils.get('cpu_percent')['values']:
            cpu_percent_values = cpu_utils.get('cpu_percent')['values']
            # Reverse to get chronological order
            cpu_percent_values.reverse()
        
        serverless_costs = calculate_serverless_cost_estimate(
            vcpu, avg_cpu_util, min_cpu_util, max_cpu_util, 
            dcu_price_per_hour, cpu_percent_values
        )
        
        serverless_cost_method1 = serverless_costs['method1_cost']
        serverless_cost_method2 = serverless_costs['method2_cost']
        min_dcu_baseline = serverless_costs['min_dcu_baseline']
        avg_dcu = serverless_costs['avg_dcu']
        
        # Calculate cost savings for both methods
        cost_savings_method1 = round(od_monthly_cost - serverless_cost_method1, 2)
        savings_percentage_method1 = round((cost_savings_method1 / od_monthly_cost * 100), 1) if od_monthly_cost > 0 else 0
        
        cost_savings_method2 = round(od_monthly_cost - serverless_cost_method2, 2)
        savings_percentage_method2 = round((cost_savings_method2 / od_monthly_cost * 100), 1) if od_monthly_cost > 0 else 0
        
        # Cost recommendation based on better savings
        recommendation_method1 = "Serverless" if cost_savings_method1 > 0 else "On-Demand"
        recommendation_method2 = "Serverless" if cost_savings_method2 > 0 else "On-Demand"
        
        # Get time range
        first_time = datetime.utcnow() - timedelta(days=30)
        last_time = datetime.utcnow()
        
        avg_cpu_list.append(avg_cpu_util)
        
        # Return result - includes both calculation methods
        result_summary = (
            f"{account_id},{docdb_region},{instance_id},{cluster_id},{engine},{engine_version},"
            f"{instance_class},{vcpu},{avg_cpu_util},{min_cpu_util},{max_cpu_util},"
            f"{first_time},{last_time},{od_monthly_cost},{storage_used_gb},{dcu_price_per_hour},{min_dcu_baseline},"
            f"{serverless_cost_method1},{cost_savings_method1},{savings_percentage_method1},{recommendation_method1},"
            f"{serverless_cost_method2},{cost_savings_method2},{savings_percentage_method2},{recommendation_method2}"
        )
        
        result_chart = (
            f"{docdb_region},{instance_id},{cluster_id},{engine},{engine_version},{instance_class},"
            f"{vcpu},{avg_cpu_util},{min_cpu_util},{max_cpu_util},{first_time},"
            f"{last_time},{od_price_per_unit},{od_monthly_cost},{dcu_price_per_hour},"
            f"{min_dcu_baseline},{serverless_cost_method1},{serverless_cost_method2},"
            f"{cost_savings_method1},{cost_savings_method2},{savings_percentage_method1},"
            f"{savings_percentage_method2},{recommendation_method1},{recommendation_method2},"
            f"{storage_used_gb}"
        )
        
        # Return result
        return (avg_cpu_util, result_summary, result_chart)
    
    except Exception as e:
        logging.error(f"Error processing instance {instance_id}: {str(e)}")
        return None
def main():
    """
    Main function: Get DocumentDB instance information and generate cost comparison report
    """
    output_result = []
    avg_cpu_list = []
    output_result_chart = []
    
    # Create DocumentDB client
    docdb = boto3.client('docdb')
    marker = None
    all_instances = []

    print("Retrieving DocumentDB instance list...")
    
    # Get all DocumentDB instances
    while True:
        try:
            if marker:
                instance_response = docdb.describe_db_instances(Marker=marker)
            else:
                instance_response = docdb.describe_db_instances()
            
            instances = instance_response['DBInstances']
            # Filter DocumentDB instances
            filtered = [
                instance for instance in instances
                if instance['Engine'] == 'docdb'
            ]
            all_instances.extend(filtered)
            
            if 'Marker' in instance_response:
                marker = instance_response['Marker']
            else:
                break
        except Exception as e:
            logging.error(f"Error retrieving DocumentDB instance list: {str(e)}")
            break

    if not all_instances:
        print("No DocumentDB instances found in the specified region")
        return

    print(f"Found {len(all_instances)} DocumentDB instances")

    # Get DocumentDB pricing information
    print("Retrieving DocumentDB pricing information...")
    pricing_get_products_optimized(docdb_region)

    # Get DCU pricing
    print("Retrieving DocumentDB Serverless DCU pricing...")
    dcu_price_per_hour = get_docdb_dcu_price(docdb_region)
    print(f"DCU price: ${dcu_price_per_hour} per DCU-Hr")

    # Process instances in batches
    batch_size = 2
    total_instances = len(all_instances)
    total_batches = (total_instances + batch_size - 1) // batch_size

    print(f"######## Total batches: {total_batches}, total instances: {total_instances} ########")
    
    # Process each batch of instances
    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min(start_idx + batch_size, total_instances)
        batch_instances = all_instances[start_idx:end_idx]
        print(f"######## Processing batch {batch_idx+1}/{total_batches}, instances {start_idx+1} ~ {end_idx}/{total_instances} ########")
        
        # Use concurrent processing for current batch instances
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            tasks = [
                executor.submit(process_instance, instance, total_instances, dcu_price_per_hour)
                for instance in batch_instances
            ]
            
            # Wait for all tasks to complete and process results
            for future in concurrent.futures.as_completed(tasks):
                try:
                    result = future.result()
                    if result:
                        results.append(result)
                except Exception as e:
                    print(f"Error processing instance: {e}")
        
        print(f"Batch {batch_idx+1} completed, successfully processed {len(results)} instances")

        # Process data
        for result in results:
            avg_cpu_list.append(result[0])
            output_result.append(result[1])
            output_result_chart.append(result[2])
    
    if not output_result:
        print("No instances were successfully processed")
        return
            
    # Generate report
    output_column = (
        "account_id,region,instance_id,cluster_id,engine,engine_version,instance_type,vcpu,"
        "CPU Avg Util%,CPU Min Util%,CPU Max Util%,StartTime,EndTime,"
        "OnDemand Monthly Cost,Storage Used(GB),DCU Price/Hr,Min DCU Baseline,"
        "Serverless Cost Method1,Cost Savings Method1,Savings % Method1,Recommendation Method1,"
        "Serverless Cost Method2,Cost Savings Method2,Savings % Method2,Recommendation Method2"
    )
    
    logging.info("DocumentDB cost comparison evaluation results below, recommend copying and pasting to Excel for viewing")
    logging.info(f"{output_column}")

    # Sort by cost savings (Method 2), show top 8
    top_savings_output_result = sorted(
        output_result_chart, 
        key=lambda x: float(x.split(',')[18]), 
        reverse=True
    )[:8]
    
    # Create charts
    create_cpu_usage_distribution_chart(
        count_cpu_usage_distribution(avg_cpu_list), 
        myworksheet, 
        myworkbook
    )
    
    if top_savings_output_result:
        create_cost_comparison_chart(top_savings_output_result, myworksheet, myworkbook)
    
    # Output results to log
    for line in output_result:
        logging.info(line)

    # Create detailed data worksheet
    myworksheet_detail = myworkbook.create_sheet("Detail")
    column_names = output_column.split(",")
    df = pd.DataFrame([line.split(",") for line in output_result], columns=column_names)
    myworksheet_detail.append(column_names)
    for row in df.itertuples(index=False):
        myworksheet_detail.append(row)

    # Create cost comparison summary worksheet
    myworksheet_summary = myworkbook.create_sheet("Cost Summary")
    summary_columns = [
        "Instance ID", "Cluster ID", "Instance Type", "vCPU", "Avg CPU %", "Min DCU",
        "On-Demand Cost", "Serverless Cost (Method1)", "Serverless Cost (Method2)", 
        "Savings (Method1)", "Savings % (Method1)", "Savings (Method2)", "Savings % (Method2)",
        "Recommendation (Method1)", "Recommendation (Method2)"
    ]
    myworksheet_summary.append(summary_columns)
    
    for line in output_result:
        cells = line.split(",")
        summary_row = [
            cells[2],   # instance_id
            cells[3],   # cluster_id
            cells[6],   # instance_type
            cells[7],   # vcpu
            cells[8],   # avg_cpu_util
            cells[16],  # min_dcu_baseline
            f"${float(cells[13]):.2f}",  # od_monthly_cost
            f"${float(cells[17]):.2f}",  # serverless_cost_method1
            f"${float(cells[21]):.2f}",  # serverless_cost_method2
            f"${float(cells[18]):.2f}",  # cost_savings_method1
            f"{float(cells[19]):.1f}%",  # savings_percentage_method1
            f"${float(cells[22]):.2f}",  # cost_savings_method2
            f"{float(cells[23]):.1f}%",  # savings_percentage_method2
            cells[20],  # recommendation_method1
            cells[24]   # recommendation_method2
        ]
        myworksheet_summary.append(summary_row)

    # Save results to xlsx file
    myworkbook.save("docdb_cost_comparison_report.xlsx")
    print(f"\nReport saved to docdb_cost_comparison_report.xlsx")
    
    # Output summary statistics
    total_od_cost = sum(float(line.split(',')[13]) for line in output_result)
    total_serverless_cost_method1 = sum(float(line.split(',')[17]) for line in output_result)
    total_serverless_cost_method2 = sum(float(line.split(',')[21]) for line in output_result)
    total_savings_method1 = total_od_cost - total_serverless_cost_method1
    total_savings_method2 = total_od_cost - total_serverless_cost_method2
    
    print(f"\n=== Cost Comparison Summary ===")
    print(f"Total On-Demand monthly cost: ${total_od_cost:.2f}")
    print(f"Total Serverless monthly cost (Method 1): ${total_serverless_cost_method1:.2f}")
    print(f"Total Serverless monthly cost (Method 2): ${total_serverless_cost_method2:.2f}")
    print(f"Total savings amount (Method 1): ${total_savings_method1:.2f}")
    print(f"Total savings percentage (Method 1): {(total_savings_method1/total_od_cost*100):.1f}%")
    print(f"Total savings amount (Method 2): ${total_savings_method2:.2f}")
    print(f"Total savings percentage (Method 2): {(total_savings_method2/total_od_cost*100):.1f}%")
    
    # Recommendation statistics
    method1_serverless_recommended = sum(1 for line in output_result if line.split(',')[20] == 'Serverless')
    method1_od_recommended = len(output_result) - method1_serverless_recommended
    method2_serverless_recommended = sum(1 for line in output_result if line.split(',')[24] == 'Serverless')
    method2_od_recommended = len(output_result) - method2_serverless_recommended
    
    print(f"\n=== Recommendation Statistics ===")
    print(f"Method 1 - Instances recommended for Serverless: {method1_serverless_recommended}")
    print(f"Method 1 - Instances recommended for On-Demand: {method1_od_recommended}")
    print(f"Method 2 - Instances recommended for Serverless: {method2_serverless_recommended}")
    print(f"Method 2 - Instances recommended for On-Demand: {method2_od_recommended}")

if __name__ == "__main__":
    try:
        main()
        print("\nProcessing complete!")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        logging.error(traceback.format_exc())
        logging.info("Please contact the AWS team for processing.")
        raise
