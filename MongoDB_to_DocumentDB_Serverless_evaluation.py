#!/usr/bin/env python3
"""
MongoDB to DocumentDB Serverless Migration Evaluation Tool
This tool evaluates the cost and performance impact of migrating self-hosted MongoDB 
on EC2 instances to Amazon DocumentDB Serverless.

Key Features:
- Identifies MongoDB EC2 instances by tag: docDB-serverless-eva=true
- Analyzes CPU utilization and EC2 costs (instance + storage)
- Compares with DocumentDB Serverless costs using real-time pricing
- Generates comprehensive Excel reports with visualizations
"""

import boto3
import traceback
import math
import json
import os
from datetime import datetime, timedelta
import time
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import logging
import sys
from operator import itemgetter
import matplotlib.pyplot as plt
import numpy as np
import concurrent.futures
from botocore.exceptions import ClientError

# Global variables
priceList = []
counter = 0
storage_cache = {}

IMG_WIDTH = 600
IMG_HEIGHT = 400

# List of supported AWS regions
region_list = ['us-east-1','ap-northeast-1','us-east-2','us-west-1','us-west-2','ap-east-1','ap-south-1','ap-southeast-1','ap-northeast-2','ap-southeast-2','ca-central-1','eu-central-1','eu-west-1','eu-west-2','eu-west-3','eu-north-1','me-south-1','sa-east-1']

print("MongoDB to DocumentDB Serverless Migration Evaluation Tool")
print("=" * 60)
print("Please select a region by entering the corresponding number:")
for i, rg in enumerate(region_list, start=1):
    print(f"{i}. {rg}")
user_input = input(f"Enter your choice (1-{len(region_list)}): ")
if user_input.isdigit() and 1 <= int(user_input) <= len(region_list):
    selected_region = region_list[int(user_input) - 1]
    print(f"You selected: {selected_region}")
else:
    print("Invalid input. Please try again.")
    sys.exit(1)

os.environ['AWS_DEFAULT_REGION'] = selected_region

# Create a new Excel workbook
myworkbook = openpyxl.Workbook()
myworksheet = myworkbook.active
myworksheet.title = "Summary"

# Get current date and time for log filename
log_filename = f"mongodb_to_docdb_evaluation_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# Configure logging
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

def aws_region_to_location(region):
    """
    Convert AWS region code to location name used in pricing API
    """
    region_to_location_map = {
        "us-east-1": "US East (N. Virginia)",
        "us-east-2": "US East (Ohio)",
        "us-west-1": "US West (N. California)",
        "us-west-2": "US West (Oregon)",
        "ap-east-1": "Asia Pacific (Hong Kong)",
        "ap-south-1": "Asia Pacific (Mumbai)",
        "ap-northeast-1": "Asia Pacific (Tokyo)",
        "ap-northeast-2": "Asia Pacific (Seoul)",
        "ap-southeast-1": "Asia Pacific (Singapore)",
        "ap-southeast-2": "Asia Pacific (Sydney)",
        "ca-central-1": "Canada (Central)",
        "eu-central-1": "EU (Frankfurt)",
        "eu-west-1": "EU (Ireland)",
        "eu-west-2": "EU (London)",
        "eu-west-3": "EU (Paris)",
        "eu-north-1": "EU (Stockholm)",
        "me-south-1": "Middle East (Bahrain)",
        "sa-east-1": "South America (São Paulo)"
    }

    if region in region_to_location_map:
        return region_to_location_map[region]
    else:
        return "Unknown location"

def get_all_pages(pricing_client, service_code, filters, max_retries=5):
    """
    Get all results from AWS Pricing API with pagination
    """
    all_results = []
    next_token = None
    retries = 0
    
    while True:
        try:
            params = {
                'ServiceCode': service_code,
                'Filters': filters
            }
            if next_token:
                params['NextToken'] = next_token
            
            response = pricing_client.get_products(**params)
            all_results.extend(response['PriceList'])
            
            if 'NextToken' in response:
                next_token = response['NextToken']
            else:
                break
                
            time.sleep(0.5)
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'Throttling':
                retries += 1
                if retries > max_retries:
                    logging.error(f"Reached maximum retry attempts: {e}")
                    break
                wait_time = 2 ** retries
                logging.warning(f"Request throttled, retrying in {wait_time} seconds ({retries}/{max_retries})")
                time.sleep(wait_time)
            else:
                logging.error(f"API request error: {e}")
                break
    
    logging.info(f"Retrieved {len(all_results)} pricing records")
    return all_results

def find_mongodb_ec2_instances(region):
    """
    Find EC2 instances tagged with docDB-serverless-eva=true
    """
    ec2_client = boto3.client('ec2', region_name=region)
    
    try:
        # Query EC2 instances with specific tag
        response = ec2_client.describe_instances(
            Filters=[
                {
                    'Name': 'tag:docDB-serverless-eva',
                    'Values': ['true']
                },
                {
                    'Name': 'instance-state-name',
                    'Values': ['running']
                }
            ]
        )
        
        instances = []
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                # Add OwnerId from reservation to instance data
                instance['OwnerId'] = reservation['OwnerId']
                instances.append(instance)
        
        logging.info(f"Found {len(instances)} MongoDB EC2 instances with evaluation tag")
        return instances
        
    except Exception as e:
        logging.error(f"Error finding MongoDB EC2 instances: {e}")
        return []

def get_ec2_pricing(region, instance_type, tenancy='Shared', operating_system='Linux'):
    """
    Get EC2 instance pricing from AWS Pricing API
    """
    pricing_client = boto3.client('pricing', region_name='us-east-1')
    location = aws_region_to_location(region)
    
    filters = [
        {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
        {'Type': 'TERM_MATCH', 'Field': 'instanceType', 'Value': instance_type},
        {'Type': 'TERM_MATCH', 'Field': 'tenancy', 'Value': tenancy},
        {'Type': 'TERM_MATCH', 'Field': 'operatingSystem', 'Value': operating_system},
        {'Type': 'TERM_MATCH', 'Field': 'preInstalledSw', 'Value': 'NA'},
        {'Type': 'TERM_MATCH', 'Field': 'capacitystatus', 'Value': 'Used'}
    ]
    
    try:
        products = get_all_pages(pricing_client, 'AmazonEC2', filters)
        
        for product_str in products:
            product = json.loads(product_str)
            
            # Get On-Demand pricing
            terms = product.get('terms', {})
            on_demand = terms.get('OnDemand', {})
            
            for offer_term_code, offer_term_data in on_demand.items():
                for price_dimension_key, price_dimension_data in offer_term_data['priceDimensions'].items():
                    price_per_unit = float(price_dimension_data['pricePerUnit']['USD'])
                    unit = price_dimension_data.get('unit', 'Hrs')
                    
                    # Get instance specifications
                    attributes = product['product']['attributes']
                    vcpu = int(attributes.get('vcpu', 0))
                    memory = attributes.get('memory', '0 GiB')
                    
                    return {
                        'price_per_hour': price_per_unit,
                        'unit': unit,
                        'vcpu': vcpu,
                        'memory': memory,
                        'instance_type': instance_type
                    }
        
        logging.error(f"No pricing found for EC2 instance type: {instance_type}")
        return None
        
    except Exception as e:
        logging.error(f"Error retrieving EC2 pricing for {instance_type}: {e}")
        return None

def get_ebs_pricing(region, volume_type='gp3'):
    """
    Get EBS storage pricing from AWS Pricing API
    """
    pricing_client = boto3.client('pricing', region_name='us-east-1')
    location = aws_region_to_location(region)
    
    # Map volume types to pricing API volume types
    volume_type_mapping = {
        'gp2': 'General Purpose',
        'gp3': 'General Purpose', 
        'io1': 'Provisioned IOPS',
        'io2': 'Provisioned IOPS',
        'st1': 'Throughput Optimized HDD',
        'sc1': 'Cold HDD',
        'standard': 'Magnetic'
    }
    
    api_volume_type = volume_type_mapping.get(volume_type, 'General Purpose')
    
    filters = [
        {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
        {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Storage'},
        {'Type': 'TERM_MATCH', 'Field': 'volumeType', 'Value': api_volume_type}
    ]
    
    try:
        products = get_all_pages(pricing_client, 'AmazonEC2', filters)
        
        for product_str in products:
            product = json.loads(product_str)
            
            # Check if this matches the specific volume type we want
            attributes = product['product']['attributes']
            volume_api_name = attributes.get('volumeApiName', '').lower()
            usage_type = attributes.get('usagetype', '').lower()
            
            # More precise matching for the specific volume type
            if (volume_type.lower() in usage_type or 
                volume_type.lower() == volume_api_name or
                (volume_type == 'gp2' and 'gp2' in usage_type) or
                (volume_type == 'gp3' and 'gp3' in usage_type)):
                
                # Get On-Demand pricing
                terms = product.get('terms', {})
                on_demand = terms.get('OnDemand', {})
                
                for offer_term_code, offer_term_data in on_demand.items():
                    for price_dimension_key, price_dimension_data in offer_term_data['priceDimensions'].items():
                        price_per_unit = float(price_dimension_data['pricePerUnit']['USD'])
                        unit = price_dimension_data.get('unit', 'GB-Mo')
                        
                        if 'GB-Mo' in unit:
                            logging.info(f"Found EBS {volume_type} pricing: ${price_per_unit} per {unit}")
                            return {
                                'price_per_gb_month': price_per_unit,
                                'unit': unit,
                                'volume_type': volume_type,
                                'usage_type': usage_type,
                                'volume_api_name': volume_api_name
                            }
        
        logging.error(f"No EBS pricing found for volume type: {volume_type}")
        return None
        
    except Exception as e:
        logging.error(f"Error retrieving EBS pricing for {volume_type}: {e}")
        return None

def get_ec2_volumes(instance_id, region):
    """
    Get EBS volumes attached to an EC2 instance
    """
    ec2_client = boto3.client('ec2', region_name=region)
    
    try:
        response = ec2_client.describe_instances(InstanceIds=[instance_id])
        
        volumes = []
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                for block_device in instance.get('BlockDeviceMappings', []):
                    if 'Ebs' in block_device:
                        volume_id = block_device['Ebs']['VolumeId']
                        
                        # Get volume details
                        volume_response = ec2_client.describe_volumes(VolumeIds=[volume_id])
                        for volume in volume_response['Volumes']:
                            volumes.append({
                                'volume_id': volume_id,
                                'size': volume['Size'],
                                'volume_type': volume['VolumeType'],
                                'iops': volume.get('Iops', 0),
                                'throughput': volume.get('Throughput', 0)
                            })
        
        return volumes
        
    except Exception as e:
        logging.error(f"Error getting volumes for instance {instance_id}: {e}")
        return []

def get_ec2_cpu_utilization(instance_id, region):
    """
    Get CPU utilization statistics for an EC2 instance over the past month
    """
    cloudwatch_client = boto3.client('cloudwatch', region_name=region)

    response = cloudwatch_client.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'cpu_avg',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/EC2',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'InstanceId',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31, 
                    'Stat': 'Average'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_min',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/EC2',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'InstanceId',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,
                    'Stat': 'Minimum'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_max',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/EC2',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'InstanceId',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,
                    'Stat': 'Maximum'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_percent',
                'Expression': 'IF(m1>0,m1)', 
                'Label': 'CPU Watch'
            },
            {
                'Id': 'm1',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/EC2',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'InstanceId',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 300,  # 5-minute resolution for detailed analysis
                    'Stat': 'Maximum',
                    'Unit': 'Percent'
                },
                'ReturnData': False
            }
        ],
      
        StartTime=(datetime.utcnow() - timedelta(days=30)).isoformat() + 'Z',
        EndTime=datetime.utcnow().isoformat() + 'Z'
    )

    if response['MetricDataResults']:
        metrics = {}
        for mdr in response['MetricDataResults']:
            metrics[mdr['Id']] = {'timestamps': mdr['Timestamps'], 'values': mdr['Values']}
        return metrics
    else:
        return None

def get_docdb_dcu_price(region):
    """
    Get DocumentDB Serverless DCU unit price
    """
    pricing_client = boto3.client('pricing', region_name='us-east-1')
    location = aws_region_to_location(region)
    
    logging.info(f"Querying real-time DCU price for {region} ({location})...")
    
    filters = [
        {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location}
    ]
    
    try:
        products = get_all_pages(pricing_client, 'AmazonDocDB', filters)
        price_list = [json.loads(product) for product in products]
        
        logging.info(f"Retrieved {len(price_list)} DocumentDB pricing records")
        
        # Search for Serverless-related pricing
        serverless_prices = []
        
        for price in price_list:
            attributes = price['product']['attributes']
            
            usage_type = attributes.get('usagetype', '').lower()
            product_family = attributes.get('productFamily', '').lower()
            instance_type = attributes.get('instanceType', '').lower()
            description = attributes.get('description', '').lower()
            
            if ('serverless' in usage_type or 
                'serverless' in product_family or 
                'serverless' in instance_type or
                'serverless' in description):
                
                terms = price.get('terms', {})
                on_demand = terms.get('OnDemand', {})
                
                for offer_term_code, offer_term_data in on_demand.items():
                    for price_dimension_key, price_dimension_data in offer_term_data['priceDimensions'].items():
                        price_per_unit = float(price_dimension_data['pricePerUnit']['USD'])
                        unit = price_dimension_data.get('unit', 'Unknown')
                        
                        serverless_prices.append({
                            'price': price_per_unit,
                            'unit': unit,
                            'usage_type': usage_type,
                            'product_family': product_family,
                            'description': price_dimension_data.get('description', 'N/A')
                        })
        
        if serverless_prices:
            dcu_candidates = [p for p in serverless_prices if 'dcu' in p['unit'].lower() or 'capacity' in p['description'].lower()]
            
            if dcu_candidates:
                best_price = dcu_candidates[0]['price']
                logging.info(f"Found Serverless DCU price: ${best_price} per {dcu_candidates[0]['unit']}")
                return best_price
            else:
                best_price = serverless_prices[0]['price']
                logging.info(f"Found Serverless price: ${best_price} per {serverless_prices[0]['unit']}")
                return best_price
        
        # If no serverless pricing found, search for DCU-related pricing
        logging.warning("No DocumentDB Serverless pricing found, searching for DCU-related pricing...")
        
        for price in price_list:
            attributes = price['product']['attributes']
            
            usage_type = attributes.get('usagetype', '').lower()
            product_family = attributes.get('productFamily', '').lower()
            description = attributes.get('description', '').lower()
            
            if ('dcu' in usage_type or 
                'dcu' in description or
                'capacity' in description or
                'io-optimized' in description):
                
                terms = price.get('terms', {})
                on_demand = terms.get('OnDemand', {})
                
                for offer_term_code, offer_term_data in on_demand.items():
                    for price_dimension_key, price_dimension_data in offer_term_data['priceDimensions'].items():
                        price_per_unit = float(price_dimension_data['pricePerUnit']['USD'])
                        unit = price_dimension_data.get('unit', 'Unknown')
                        
                        if 'io-optimized' in description or 'dcu' in unit.lower():
                            logging.info(f"Found DCU-related price: ${price_per_unit} per {unit}")
                            return price_per_unit
        
        logging.error("No DocumentDB Serverless DCU pricing found in API")
        raise Exception("Unable to retrieve DocumentDB Serverless DCU pricing from AWS Pricing API")
        
    except Exception as e:
        logging.error(f"Error retrieving DCU price: {e}")
        raise Exception(f"Failed to retrieve DocumentDB Serverless DCU pricing: {e}")

def calculate_serverless_cost_estimate(vcpu, avg_cpu_util, min_cpu_util, max_cpu_util, dcu_price_per_hour, cpu_percent_data=None):
    """
    Calculate estimated DocumentDB Serverless cost for MongoDB workload
    Uses conservative estimation based on CPU utilization patterns
    """
    
    # Method 1: Simple average-based calculation
    # Estimate DCU based on CPU utilization (conservative multiplier for MongoDB workloads)
    estimated_dcu_per_instance = vcpu * (avg_cpu_util / 100.0)
    min_dcu_per_instance = max(0.5, estimated_dcu_per_instance)  # Minimum 0.5 DCU
    
    # Monthly cost calculation (730 hours per month)
    serverless_cost_method1 = min_dcu_per_instance * 730 * dcu_price_per_hour
    
    logging.info(f"Method 1 - Simple calculation: vCPU={vcpu}, avg_cpu_util={avg_cpu_util}%, "
                f"estimated_dcu={min_dcu_per_instance}, "
                f"monthly cost=${serverless_cost_method1:.2f}")
    
    # Method 2: Detailed calculation with baseline + burst
    # Calculate baseline DCU (minimum sustained load)
    baseline_dcu = max(0.5, vcpu * (min_cpu_util / 100.0))
    
    # Base cost (baseline DCU)
    base_cost = baseline_dcu * dcu_price_per_hour * 730
    
    # Calculate burst cost if detailed CPU data is available
    burst_cost = 0
    if cpu_percent_data and len(cpu_percent_data) > 0:
        burst_threshold = (avg_cpu_util + min_cpu_util) / 2
        
        logging.info(f"Calculating burst cost with threshold: {burst_threshold}%")
        
        burst_count = 0
        total_burst_cost = 0
        
        for cpu_value in cpu_percent_data:
            if cpu_value > burst_threshold:
                required_dcu = vcpu * (cpu_value / 100.0)
                additional_dcu = max(0, required_dcu - baseline_dcu)
                
                # Additional cost charged per 5-minute period
                additional_cost = additional_dcu * dcu_price_per_hour / 12  # 5 minutes = 1/12 hour
                total_burst_cost += additional_cost
                burst_count += 1
        
        burst_cost = total_burst_cost
        
        logging.info(f"Burst scaling: {burst_count} periods exceeded threshold, "
                    f"additional cost=${total_burst_cost:.2f}")
    
    serverless_cost_method2 = base_cost + burst_cost
    
    logging.info(f"Method 2 - Detailed calculation: baseline_dcu={baseline_dcu}, "
                f"base_cost=${base_cost:.2f}, burst_cost=${burst_cost:.2f}, "
                f"total monthly cost=${serverless_cost_method2:.2f}")
    
    return {
        'method1_cost': round(serverless_cost_method1, 2),
        'method2_cost': round(serverless_cost_method2, 2),
        'baseline_dcu': baseline_dcu,
        'estimated_dcu': min_dcu_per_instance
    }

def count_cpu_usage_distribution(cpu_usage_data):
    """
    Count CPU utilization distribution for MongoDB instances
    """
    usage_ranges = [
        ('0% - 10%', 0, 10),
        ('10% - 20%', 10, 20),
        ('20% - 30%', 20, 30),
        ('30% - 50%', 30, 50),
        ('50% and above', 50, 100)
    ]

    usage_counts = [0] * len(usage_ranges)

    for usage in cpu_usage_data:
        for i, (_, min_range, max_range) in enumerate(usage_ranges):
            if min_range <= usage < max_range:
                usage_counts[i] += 1
                break

    result = [['CPU Usage Range', 'Instance Count']]
    result.extend([[range_name, count] for range_name, count in zip(
        [range_name for range_name, _, _ in usage_ranges], usage_counts
    )])
    return result

def create_cpu_usage_distribution_chart(data, worksheet, workbook):
    """
    Create CPU utilization distribution pie chart for MongoDB instances
    """
    for row in data:
        worksheet.append(row)

    labels = [row[0] for row in data[1:] if float(row[1]) > 0]
    values = [float(row[1]) for row in data[1:] if float(row[1]) > 0]

    fig, ax = plt.subplots(figsize=(8, 8))
    patches, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%')
    
    for t in texts:
        t.set_size('smaller')
        t.set_color('black')

    for t in autotexts:
        t.set_size('smaller')
        t.set_color('white')

    ax.axis('equal')
    ax.set_title("MongoDB EC2 Instance Count by CPU Avg Utilization")

    plt.savefig("mongodb_cpu_usage_pie.jpg", dpi=300)

    img = Image("mongodb_cpu_usage_pie.jpg")
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    worksheet.add_image(img, "A16")

def create_cost_comparison_chart(data, worksheet, workbook):
    """
    Create cost comparison chart for MongoDB to DocumentDB Serverless migration
    """
    id = 2
    worksheet.cell(row=1, column=12, value='instance')
    worksheet.cell(row=1, column=13, value='mongodb cost')
    worksheet.cell(row=1, column=14, value='serverless cost method1')
    worksheet.cell(row=1, column=15, value='serverless cost method2')
    worksheet.cell(row=1, column=16, value='cost savings method1')
    worksheet.cell(row=1, column=17, value='cost savings method2')
    
    cluster_cost_data = []
    
    for row in data:
        cells = row.split(',')
        # Select instance name, mongodb cost, serverless costs (both methods)
        selected_columns = [3, 16, 19, 23, 20, 24]  # instance_id, mongodb_cost, serverless_method1, serverless_method2, savings_method1, savings_method2
        selected_cell = [cells[i] for i in selected_columns]
        
        for col, value in enumerate(selected_cell, start=1):
            if col > 1:  # All except instance name are numeric
                value = float(value)
            worksheet.cell(row=id, column=12 + col - 1, value=value)
        
        cluster_cost_data.extend(selected_cell)
        id = id + 1

    # Create cost comparison bar chart
    instances = []
    mongodb_costs = []
    serverless_costs_method1 = []
    serverless_costs_method2 = []
    savings_method1 = []
    savings_method2 = []

    for i in range(0, len(cluster_cost_data), 6):
        instances.append(cluster_cost_data[i])
        mongodb_costs.append(float(cluster_cost_data[i+1]))
        serverless_costs_method1.append(float(cluster_cost_data[i+2]))
        serverless_costs_method2.append(float(cluster_cost_data[i+3]))
        savings_method1.append(float(cluster_cost_data[i+4]))
        savings_method2.append(float(cluster_cost_data[i+5]))

    img = Image(create_cost_comparison_bar_chart(instances, mongodb_costs, serverless_costs_method1, serverless_costs_method2, savings_method1, savings_method2))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    worksheet.add_image(img, "L16")

def create_cost_comparison_bar_chart(instances, mongodb_costs, serverless_costs_method1, serverless_costs_method2, savings_method1, savings_method2):
    """
    Create cost comparison bar chart JPG image for MongoDB to DocumentDB Serverless
    """
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # First chart: cost comparison - Method 1
    bar_width = 0.35
    x = np.arange(len(instances))
    
    bars1 = ax1.bar(x - bar_width/2, mongodb_costs, bar_width, label='MongoDB on EC2', color='skyblue')
    bars2 = ax1.bar(x + bar_width/2, serverless_costs_method1, bar_width, label='DocumentDB Serverless (Method 1)', color='lightcoral')

    # Add value labels
    for i, (mongo, docdb) in enumerate(zip(mongodb_costs, serverless_costs_method1)):
        ax1.text(x[i] - bar_width/2, mongo, f'${mongo:.0f}', ha='center', va='bottom', fontsize=8)
        ax1.text(x[i] + bar_width/2, docdb, f'${docdb:.0f}', ha='center', va='bottom', fontsize=8)

    ax1.set_title("Cost Comparison: MongoDB on EC2 vs DocumentDB Serverless (Method 1)", fontsize=12)
    ax1.set_xlabel("Instance", fontsize=10)
    ax1.set_ylabel("Monthly Cost (USD)", fontsize=10)
    ax1.set_xticks(x)
    ax1.set_xticklabels(instances, rotation=20, fontsize=8)
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Second chart: cost comparison - Method 2
    bars3 = ax2.bar(x - bar_width/2, mongodb_costs, bar_width, label='MongoDB on EC2', color='skyblue')
    bars4 = ax2.bar(x + bar_width/2, serverless_costs_method2, bar_width, label='DocumentDB Serverless (Method 2)', color='orange')

    # Add value labels
    for i, (mongo, docdb) in enumerate(zip(mongodb_costs, serverless_costs_method2)):
        ax2.text(x[i] - bar_width/2, mongo, f'${mongo:.0f}', ha='center', va='bottom', fontsize=8)
        ax2.text(x[i] + bar_width/2, docdb, f'${docdb:.0f}', ha='center', va='bottom', fontsize=8)

    ax2.set_title("Cost Comparison: MongoDB on EC2 vs DocumentDB Serverless (Method 2)", fontsize=12)
    ax2.set_xlabel("Instance", fontsize=10)
    ax2.set_ylabel("Monthly Cost (USD)", fontsize=10)
    ax2.set_xticks(x)
    ax2.set_xticklabels(instances, rotation=20, fontsize=8)
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    # Third chart: savings comparison - Method 1
    colors1 = ['green' if s > 0 else 'red' for s in savings_method1]
    bars5 = ax3.bar(instances, savings_method1, color=colors1, alpha=0.7)
    
    # Add value labels
    for i, s in enumerate(savings_method1):
        ax3.text(i, s, f'${s:.0f}', ha='center', va='bottom' if s > 0 else 'top', fontsize=8)
    
    ax3.set_title("Cost Savings - Method 1 (Positive = DocumentDB Serverless Cheaper)", fontsize=12)
    ax3.set_xlabel("Instance", fontsize=10)
    ax3.set_ylabel("Savings (USD)", fontsize=10)
    ax3.set_xticklabels(instances, rotation=20, fontsize=8)
    ax3.axhline(y=0, color='black', linestyle='-', alpha=0.3)
    ax3.grid(True, alpha=0.3)
    
    # Fourth chart: savings comparison - Method 2
    colors2 = ['green' if s > 0 else 'red' for s in savings_method2]
    bars6 = ax4.bar(instances, savings_method2, color=colors2, alpha=0.7)
    
    # Add value labels
    for i, s in enumerate(savings_method2):
        ax4.text(i, s, f'${s:.0f}', ha='center', va='bottom' if s > 0 else 'top', fontsize=8)
    
    ax4.set_title("Cost Savings - Method 2 (Positive = DocumentDB Serverless Cheaper)", fontsize=12)
    ax4.set_xlabel("Instance", fontsize=10)
    ax4.set_ylabel("Savings (USD)", fontsize=10)
    ax4.set_xticklabels(instances, rotation=20, fontsize=8)
    ax4.axhline(y=0, color='black', linestyle='-', alpha=0.3)
    ax4.grid(True, alpha=0.3)

    plt.tight_layout()
    
    file_name = "mongodb_to_docdb_cost_comparison_chart.jpg"
    plt.savefig(file_name, dpi=300, bbox_inches='tight')
    plt.close()
    return file_name

def update_progress(current_step, total_steps):
    """
    Update progress bar
    """
    progress = current_step / total_steps * 100
    bar_length = 30
    block = int(round(bar_length * progress / 100))
    text = "\rProcessing: [{0}] {1}%".format("#" * block + "-" * (bar_length - block), round(progress, 2))
    sys.stdout.write(text)
    sys.stdout.flush()

def process_mongodb_instance(instance, instance_count, dcu_price_per_hour, ec2_pricing_cache, ebs_pricing_cache):
    """
    Process a single MongoDB EC2 instance and calculate migration costs
    """
    global counter
    avg_cpu_list = []
    counter += 1
    update_progress(counter, instance_count)
    logging.info("-----------------------")
    
    instance_id = instance['InstanceId']
    instance_type = instance['InstanceType']
    availability_zone = instance['Placement']['AvailabilityZone']
    account_id = instance.get('OwnerId', 'Unknown')
    
    # Get instance name from tags
    instance_name = instance_id
    for tag in instance.get('Tags', []):
        if tag['Key'] == 'Name':
            instance_name = tag['Value']
            break
    
    try:
        # Get EC2 pricing information (with caching)
        if instance_type not in ec2_pricing_cache:
            ec2_pricing = get_ec2_pricing(selected_region, instance_type)
            if not ec2_pricing:
                logging.error(f"Could not get pricing for instance type {instance_type}")
                return None
            ec2_pricing_cache[instance_type] = ec2_pricing
        else:
            ec2_pricing = ec2_pricing_cache[instance_type]
        
        vcpu = ec2_pricing['vcpu']
        memory = ec2_pricing['memory']
        ec2_hourly_cost = ec2_pricing['price_per_hour']
        
        # Get EBS volumes and calculate storage costs
        volumes = get_ec2_volumes(instance_id, selected_region)
        total_storage_gb = 0
        total_storage_cost = 0
        
        for volume in volumes:
            volume_type = volume['volume_type']
            volume_size = volume['size']
            total_storage_gb += volume_size
            
            # Get EBS pricing (with caching)
            cache_key = f"{volume_type}"
            if cache_key not in ebs_pricing_cache:
                ebs_pricing = get_ebs_pricing(selected_region, volume_type)
                if ebs_pricing:
                    ebs_pricing_cache[cache_key] = ebs_pricing
                else:
                    logging.warning(f"Could not get EBS pricing for {volume_type}, using default gp3")
                    ebs_pricing = get_ebs_pricing(selected_region, 'gp3')
                    ebs_pricing_cache[cache_key] = ebs_pricing
            else:
                ebs_pricing = ebs_pricing_cache[cache_key]
            
            if ebs_pricing:
                volume_monthly_cost = volume_size * ebs_pricing['price_per_gb_month']
                total_storage_cost += volume_monthly_cost
        
        # Get CPU utilization for this instance
        cpu_utils = get_ec2_cpu_utilization(instance_id, selected_region)
        if cpu_utils and cpu_utils.get('cpu_avg') and cpu_utils.get('cpu_avg')['values']:
            avg_cpu_util = math.ceil(cpu_utils.get('cpu_avg')['values'][0])
            min_cpu_util = math.ceil(cpu_utils.get('cpu_min')['values'][0])
            max_cpu_util = math.ceil(cpu_utils.get('cpu_max')['values'][0])
        else:
            avg_cpu_util = 0
            min_cpu_util = 0
            max_cpu_util = 0
            logging.warning(f"Instance {instance_id} has no CPU utilization data")
        
        # Calculate MongoDB total monthly cost (EC2 + Storage)
        ec2_monthly_cost = round(730 * ec2_hourly_cost, 2)
        mongodb_total_cost = round(ec2_monthly_cost + total_storage_cost, 2)
        
        # Calculate DocumentDB Serverless cost using dual methods
        cpu_percent_values = None
        if cpu_utils and cpu_utils.get('cpu_percent') and cpu_utils.get('cpu_percent')['values']:
            cpu_percent_values = cpu_utils.get('cpu_percent')['values']
            cpu_percent_values.reverse()  # Get chronological order
        
        serverless_costs = calculate_serverless_cost_estimate(
            vcpu, avg_cpu_util, min_cpu_util, max_cpu_util, 
            dcu_price_per_hour, cpu_percent_values
        )
        
        serverless_cost_method1 = serverless_costs['method1_cost']
        serverless_cost_method2 = serverless_costs['method2_cost']
        baseline_dcu = serverless_costs['baseline_dcu']
        estimated_dcu = serverless_costs['estimated_dcu']
        
        # Calculate cost savings for both methods
        cost_savings_method1 = round(mongodb_total_cost - serverless_cost_method1, 2)
        savings_percentage_method1 = round((cost_savings_method1 / mongodb_total_cost * 100), 1) if mongodb_total_cost > 0 else 0
        
        cost_savings_method2 = round(mongodb_total_cost - serverless_cost_method2, 2)
        savings_percentage_method2 = round((cost_savings_method2 / mongodb_total_cost * 100), 1) if mongodb_total_cost > 0 else 0
        
        # Migration recommendation based on cost savings
        recommendation_method1 = "DocumentDB Serverless" if cost_savings_method1 > 0 else "Keep MongoDB on EC2"
        recommendation_method2 = "DocumentDB Serverless" if cost_savings_method2 > 0 else "Keep MongoDB on EC2"
        
        # Get time range
        first_time = datetime.utcnow() - timedelta(days=30)
        last_time = datetime.utcnow()
        
        avg_cpu_list.append(avg_cpu_util)
        
        # Return result - includes both calculation methods
        result_summary = (
            f"{account_id},{selected_region},{availability_zone},{instance_id},{instance_name},{instance_type},{vcpu},{memory},"
            f"{avg_cpu_util},{min_cpu_util},{max_cpu_util},{total_storage_gb},{first_time},{last_time},"
            f"{ec2_monthly_cost},{total_storage_cost},{mongodb_total_cost},{dcu_price_per_hour},"
            f"{baseline_dcu},{serverless_cost_method1},{cost_savings_method1},{savings_percentage_method1},{recommendation_method1},"
            f"{serverless_cost_method2},{cost_savings_method2},{savings_percentage_method2},{recommendation_method2}"
        )
        
        result_chart = (
            f"{selected_region},{availability_zone},{instance_id},{instance_name},{instance_type},{vcpu},{memory},"
            f"{avg_cpu_util},{min_cpu_util},{max_cpu_util},{total_storage_gb},{first_time},"
            f"{last_time},{ec2_hourly_cost},{ec2_monthly_cost},{total_storage_cost},{mongodb_total_cost},"
            f"{dcu_price_per_hour},{baseline_dcu},{serverless_cost_method1},{serverless_cost_method2},"
            f"{cost_savings_method1},{cost_savings_method2},{savings_percentage_method1},"
            f"{savings_percentage_method2},{recommendation_method1},{recommendation_method2}"
        )
        
        return (avg_cpu_util, result_summary, result_chart)
    
    except Exception as e:
        logging.error(f"Error processing instance {instance_id}: {str(e)}")
        return None

def main():
    """
    Main function: Find MongoDB EC2 instances and generate migration evaluation report
    """
    output_result = []
    avg_cpu_list = []
    output_result_chart = []
    ec2_pricing_cache = {}
    ebs_pricing_cache = {}
    
    print("Finding MongoDB EC2 instances with evaluation tag...")
    
    # Find MongoDB EC2 instances with the specified tag
    mongodb_instances = find_mongodb_ec2_instances(selected_region)
    
    if not mongodb_instances:
        print("No MongoDB EC2 instances found with tag 'docDB-serverless-eva=true' in the specified region")
        print("\nTo use this tool, please tag your MongoDB EC2 instances with:")
        print("Key: docDB-serverless-eva")
        print("Value: true")
        return

    print(f"Found {len(mongodb_instances)} MongoDB EC2 instances for evaluation")

    # Get DocumentDB Serverless DCU pricing
    print("Retrieving DocumentDB Serverless DCU pricing...")
    dcu_price_per_hour = get_docdb_dcu_price(selected_region)
    print(f"DCU price: ${dcu_price_per_hour} per DCU-Hr")

    # Process instances in batches
    batch_size = 2
    total_instances = len(mongodb_instances)
    total_batches = (total_instances + batch_size - 1) // batch_size

    print(f"######## Total batches: {total_batches}, total instances: {total_instances} ########")
    
    # Process each batch of instances
    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min(start_idx + batch_size, total_instances)
        batch_instances = mongodb_instances[start_idx:end_idx]
        print(f"######## Processing batch {batch_idx+1}/{total_batches}, instances {start_idx+1} ~ {end_idx}/{total_instances} ########")
        
        # Use concurrent processing for current batch instances
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            tasks = [
                executor.submit(process_mongodb_instance, instance, total_instances, dcu_price_per_hour, ec2_pricing_cache, ebs_pricing_cache)
                for instance in batch_instances
            ]
            
            # Wait for all tasks to complete and process results
            for future in concurrent.futures.as_completed(tasks):
                try:
                    result = future.result()
                    if result:
                        results.append(result)
                except Exception as e:
                    print(f"Error processing instance: {e}")
        
        print(f"Batch {batch_idx+1} completed, successfully processed {len(results)} instances")

        # Process data
        for result in results:
            avg_cpu_list.append(result[0])
            output_result.append(result[1])
            output_result_chart.append(result[2])
    
    if not output_result:
        print("No instances were successfully processed")
        return
            
    # Generate report
    output_column = (
        "account_id,region,Availability Zone,instance_id,instance_name,instance_type,vcpu,memory,"
        "CPU Avg Util%,CPU Min Util%,CPU Max Util%,Total Storage GB,StartTime,EndTime,"
        "EC2 Monthly Cost,Storage Monthly Cost,MongoDB Total Cost,DCU Price/Hr,"
        "Baseline DCU,Serverless Cost Method1,Cost Savings Method1,Savings % Method1,Recommendation Method1,"
        "Serverless Cost Method2,Cost Savings Method2,Savings % Method2,Recommendation Method2"
    )
    
    logging.info("MongoDB to DocumentDB Serverless migration evaluation results:")
    logging.info(f"{output_column}")

    # Sort by cost savings (Method 2), show top 8
    top_savings_output_result = sorted(
        output_result_chart, 
        key=lambda x: float(x.split(',')[21]), 
        reverse=True
    )[:8]
    
    # Create charts
    create_cpu_usage_distribution_chart(
        count_cpu_usage_distribution(avg_cpu_list), 
        myworksheet, 
        myworkbook
    )
    
    if top_savings_output_result:
        create_cost_comparison_chart(top_savings_output_result, myworksheet, myworkbook)
    
    # Output results to log
    for line in output_result:
        logging.info(line)

    # Create detailed data worksheet
    myworksheet_detail = myworkbook.create_sheet("Detail")
    column_names = output_column.split(",")
    df = pd.DataFrame([line.split(",") for line in output_result], columns=column_names)
    myworksheet_detail.append(column_names)
    for row in df.itertuples(index=False):
        myworksheet_detail.append(row)

    # Create cost comparison summary worksheet
    myworksheet_summary = myworkbook.create_sheet("Cost Summary")
    summary_columns = [
        "Instance ID", "Instance Name", "Instance Type", "vCPU", "Memory", "Storage GB", "Avg CPU %", "Baseline DCU",
        "MongoDB Total Cost", "DocumentDB Serverless (Method1)", "DocumentDB Serverless (Method2)", 
        "Savings (Method1)", "Savings % (Method1)", "Savings (Method2)", "Savings % (Method2)",
        "Recommendation (Method1)", "Recommendation (Method2)"
    ]
    myworksheet_summary.append(summary_columns)
    
    for line in output_result:
        cells = line.split(",")
        summary_row = [
            cells[3],   # instance_id
            cells[4],   # instance_name
            cells[5],   # instance_type
            cells[6],   # vcpu
            cells[7],   # memory
            cells[11],  # total_storage_gb
            cells[8],   # avg_cpu_util
            cells[18],  # baseline_dcu
            f"${float(cells[16]):.2f}",  # mongodb_total_cost
            f"${float(cells[19]):.2f}",  # serverless_cost_method1
            f"${float(cells[23]):.2f}",  # serverless_cost_method2
            f"${float(cells[20]):.2f}",  # cost_savings_method1
            f"{float(cells[21]):.1f}%",  # savings_percentage_method1
            f"${float(cells[24]):.2f}",  # cost_savings_method2
            f"{float(cells[25]):.1f}%",  # savings_percentage_method2
            cells[22],  # recommendation_method1
            cells[26]   # recommendation_method2
        ]
        myworksheet_summary.append(summary_row)
    # Save results to xlsx file
    myworkbook.save("mongodb_to_docdb_evaluation_report.xlsx")
    print(f"\nReport saved to mongodb_to_docdb_evaluation_report.xlsx")
    
    # Output summary statistics
    total_mongodb_cost = sum(float(line.split(',')[16]) for line in output_result)
    total_serverless_cost_method1 = sum(float(line.split(',')[19]) for line in output_result)
    total_serverless_cost_method2 = sum(float(line.split(',')[23]) for line in output_result)
    total_savings_method1 = total_mongodb_cost - total_serverless_cost_method1
    total_savings_method2 = total_mongodb_cost - total_serverless_cost_method2
    
    print(f"\n=== MongoDB to DocumentDB Serverless Migration Cost Analysis ===")
    print(f"Total MongoDB monthly cost (EC2 + Storage): ${total_mongodb_cost:.2f}")
    print(f"Total DocumentDB Serverless monthly cost (Method 1): ${total_serverless_cost_method1:.2f}")
    print(f"Total DocumentDB Serverless monthly cost (Method 2): ${total_serverless_cost_method2:.2f}")
    print(f"Total savings amount (Method 1): ${total_savings_method1:.2f}")
    print(f"Total savings percentage (Method 1): {(total_savings_method1/total_mongodb_cost*100):.1f}%")
    print(f"Total savings amount (Method 2): ${total_savings_method2:.2f}")
    print(f"Total savings percentage (Method 2): {(total_savings_method2/total_mongodb_cost*100):.1f}%")
    
    # Migration recommendation statistics
    method1_serverless_recommended = sum(1 for line in output_result if 'DocumentDB Serverless' in line.split(',')[22])
    method1_mongodb_recommended = len(output_result) - method1_serverless_recommended
    method2_serverless_recommended = sum(1 for line in output_result if 'DocumentDB Serverless' in line.split(',')[26])
    method2_mongodb_recommended = len(output_result) - method2_serverless_recommended
    
    print(f"\n=== Migration Recommendation Statistics ===")
    print(f"Method 1 - Instances recommended for DocumentDB Serverless: {method1_serverless_recommended}")
    print(f"Method 1 - Instances recommended to keep MongoDB on EC2: {method1_mongodb_recommended}")
    print(f"Method 2 - Instances recommended for DocumentDB Serverless: {method2_serverless_recommended}")
    print(f"Method 2 - Instances recommended to keep MongoDB on EC2: {method2_mongodb_recommended}")

if __name__ == "__main__":
    try:
        main()
        print("\nProcessing complete!")
        print("\nNext Steps:")
        print("1. Review the generated Excel report: mongodb_to_docdb_evaluation_report.xlsx")
        print("2. Check the log file for detailed processing information")
        print("3. Consider the migration recommendations based on your specific requirements")
        print("4. Plan a pilot migration for instances with highest cost savings potential")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        logging.error(traceback.format_exc())
        logging.info("Please contact the AWS team for assistance.")
        raise
